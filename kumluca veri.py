import pandas as pd
import schedule
import time
from datetime import datetime
import ssl
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
import sys
import shutil
import threading
import urllib.error

# --- Kaynak Yolu (EXE için) ---
def kaynak_yolu(relative_path):
    """ .exe olarak paketlendiğinde doğru dosya yolunu bulur. """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# --- Kategori Kuralları ---
KATEGORI_DOSYASI = kaynak_yolu('kategoriler.xlsx')

def kategori_kural_yukle():
    if not os.path.exists(KATEGORI_DOSYASI):
        print(f"!!! HATA: '{KATEGORI_DOSYASI}' dosyası bulunamadı!")
        return None
    try:
        df = pd.read_excel(KATEGORI_DOSYASI)
        print(f"--- BİLGİ: '{KATEGORI_DOSYASI}' başarıyla yüklendi. İçinde {len(df)} kural bulundu.")
        if 'Anahtar_Kelime' not in df.columns or 'Kategori' not in df.columns:
            print("!!! HATA: 'kategoriler.xlsx' dosyasındaki sütun başlıkları yanlış!")
            return None
        return df
    except Exception as e:
        print(f"!!! HATA: '{KATEGORI_DOSYASI}' dosyası okunurken bir hata oluştu: {e}")
        return None

kategori_df = kategori_kural_yukle()

# --- Akıllı Kategorizasyon (Değişiklik yok) ---
def normalize_turkish(text):
    if not isinstance(text, str):
        text = str(text)
    replacements = (
        ("ı", "i"), ("İ", "i"), ("ğ", "g"), ("Ğ", "g"), ("ü", "u"), ("Ü", "u"),
        ("ş", "s"), ("Ş", "s"), ("ö", "o"), ("Ö", "o"), ("ç", "c"), ("Ç", "c")
    )
    for old, new in replacements:
        text = text.replace(old, new)
    return text.lower()

def kategori_belirle(urun_adi, kurallar_df):
    if kurallar_df is None or kurallar_df.empty:
        return ''
    urun_adi_normalized = normalize_turkish(urun_adi)
    for index, satir in kurallar_df.iterrows():
        anahtar_kelime_str = str(satir['Anahtar_Kelime']).strip()
        anahtar_kelime_normalized = normalize_turkish(anahtar_kelime_str)
        if anahtar_kelime_normalized in urun_adi_normalized:
            return satir['Kategori']
    return ''

# --- Global Ayarlar (Değişiklik yok) ---
ssl._create_default_https_context = ssl._create_unverified_context
URL = "https://www.batiakdeniztv.com/kumluca-hal-fiyatlari/"
EXCEL_DOSYASI = "kumluca_hal_fiyatlari.xlsx"
YEDekLER_KLASORU = "yedekler"

is_running_lock = threading.Lock()

# --- Excel Stilleri (YAZIM HATASI DÜZELTİLDİ) ---
def excel_stillerini_uygula(dosya_yolu):
    try:
        workbook = load_workbook(dosya_yolu)
        worksheet = workbook.active
        
        RENK_MAP = {
            "Meyve": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
            "Sebze": PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
            "Yeşillik": PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid"),
        }
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ince_kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        merkezi_hizalama = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet["1:1"]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = merkezi_hizalama; cell.border = ince_kenarlik

        for row_index in range(2, worksheet.max_row + 1):
            kategori = worksheet.cell(row=row_index, column=2).value
            current_fill = RENK_MAP.get(kategori)

            for col_index in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.border = ince_kenarlik; cell.alignment = merkezi_hizalama
                if current_fill: cell.fill = current_fill
        
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = max_length + 4
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
        
        workbook.save(dosya_yolu)
        print("--- BİLGİ: [Kumluca] Gelişmiş Excel stilleri (Filtreleme dahil) başarıyla uygulandı.")
    except Exception as e:
        print(f"!!! HATA: [Kumluca] Excel stilleri uygulanırken bir hata oluştu: {e}")

# --- Ana İşlem (Kumluca'ya özel "header=0" mantığı) ---
def verileri_cek_ve_kaydet():
    if kategori_df is None:
        print("Kategorizasyon kuralları yüklenemediği için işlem durduruldu.")
        return
        
    if not is_running_lock.acquire(blocking=False):
        print("Önceki görev tamamlanmadı; bu döngü atlandı.")
        return
        
    try:
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Kumluca] Görev başladı. Veriler çekiliyor...")
        
        STANDARD_COLS_ORDER = ['Ürün Adı', 'Kategori', 'En Düşük Fiyat (TL)', 'En Yüksek Fiyat (TL)', 'Birim']
        
        # --- YENİDEN DENEME BLOĞU ---
        tablolar = None
        max_retries = 3
        for attempt in range(1, max_retries + 1):
            try:
                print(f"--- BİLGİ: [Kumluca] Veri çekiliyor (Deneme {attempt}/{max_retries}): {URL}")
                # header=0 -> Tespit script'inde veriyi gören doğru parametre buydu.
                tablolar = pd.read_html(URL, header=0) 
                print(f"--- BİLGİ: [Kumluca] Veri (Deneme {attempt}/{max_retries}) BAŞARILI. {len(tablolar)} tablo bulundu.")
                break 
            except (urllib.error.URLError, ssl.SSLError, ConnectionError, OSError) as e:
                print(f"--- UYARI: [Kumluca] AĞ HATASI (Deneme {attempt}/{max_retries}): {e}")
                if attempt < max_retries:
                    print(f"    ... 3 saniye beklenip yeniden denenecek ...")
                    time.sleep(3)
                else:
                    print(f"    ... {max_retries} deneme de başarısız oldu. Bu URL atlanıyor.")
            except Exception as e:
                print(f"--- UYARI: [Kumluca] VERİ HATASI (muhtemelen boş sayfa/veri yok): {e}")
                break
        
        if not tablolar:
            print(f"!!! HATA: [Kumluca] Veri çekme işlemi {max_retries} deneme sonunda başarısız oldu.")
            return 
        # --- YENİDEN DENEME BLOĞU SONU ---

        # --- YENİ "BİRLEŞTİR VE TEMİZLE" (header=0 metodu) ---
        
        # 1. İlk tabloyu (Tablo 0) al ve standartlaştır
        try:
            df_0 = tablolar[0]
            # Sütunları 'Ürün Adı' ve 'Fiyat' olarak yeniden adlandır
            df_0 = df_0.rename(columns={'Ürünler': 'Ürün Adı', 'Fiyat (₺/kg)': 'Fiyat'})
            # Sadece bu iki sütunu al
            df_0 = df_0[['Ürün Adı', 'Fiyat']] 
        except Exception as e:
            print(f"!!! HATA: [Kumluca] Tablo 0 (ana tablo) işlenemedi. Hata: {e}")
            return
            
        all_data_rows = [df_0] # İlk tablonun verilerini listeye ekle

        # 2. Diğer 3 tabloyu (1, 2, 3) işle
        # Bu tabloların "başlığı" aslında verinin ilk satırıdır.
        for table in tablolar[1:]:
            try:
                # Başlıkları (örn: 'Salatalık', '10.00₺') al
                header_data = list(table.columns) 
                # Bu başlığı, standart sütun isimleriyle bir DataFrame'e dönüştür
                header_df = pd.DataFrame([header_data], columns=['Ürün Adı', 'Fiyat'])
                all_data_rows.append(header_df)
                
                # Şimdi tablonun geri kalan verisini al
                data_df = table.copy()
                data_df.columns = ['Ürün Adı', 'Fiyat'] # Sütun adlarını standartlaştır
                all_data_rows.append(data_df)
            except Exception as e:
                print(f"--- UYARI: [Kumluca] Bir alt-tablo işlenirken hata (atlandı): {e}")

        # 3. Tüm parçaları (df_0 + diğerleri) birleştir
        toplam_df = pd.concat(all_data_rows, ignore_index=True)
        
        # 4. Standart formata çevir
        fiyat_df = pd.DataFrame()
        fiyat_df['Ürün Adı'] = toplam_df['Ürün Adı']
        # Sitede sadece 1 fiyat var, Min ve Max için aynı fiyatı kullan
        fiyat_df['En Düşük Fiyat (TL)'] = toplam_df['Fiyat']
        fiyat_df['En Yüksek Fiyat (TL)'] = toplam_df['Fiyat']
        fiyat_df['Birim'] = 'KG' # Sitede birim ₺/kg olarak belirtilmişti
        
        # Fiyat sütunundaki '₺' ve ' ' (boşluk) gibi kirli verileri temizle
        fiyat_df['En Düşük Fiyat (TL)'] = fiyat_df['En Düşük Fiyat (TL)'].astype(str).str.replace('₺', '', regex=False).str.strip()
        fiyat_df['En Yüksek Fiyat (TL)'] = fiyat_df['En Yüksek Fiyat (TL)'].astype(str).str.replace('₺', '', regex=False).str.strip()
        
        # --- BİRLEŞTİRME BLOĞU SONU ---

        # Kategorizasyon (Akıllı fonksiyonu kullanır)
        fiyat_df['Kategori'] = fiyat_df['Ürün Adı'].apply(lambda urun: kategori_belirle(urun, kategori_df))
        
        # Sütunları sırala
        fiyat_df = fiyat_df[STANDARD_COLS_ORDER] 
        
        print("--- BİLGİ: [Kumluca] Kategorizasyon sonrası verilerin ilk 5 satırı:")
        print(fiyat_df.head())
        
        fiyat_df.to_excel(EXCEL_DOSYASI, index=False, engine='openpyxl')
        excel_stillerini_uygula(EXCEL_DOSYASI)
        
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Kumluca] Veriler başarıyla '{EXCEL_DOSYASI}' dosyasına kaydedildi.")
        print("-" * 50)
        
    except Exception as e:
        print(f"!!! HATA: [Kumluca] Ana işlem sırasında beklenmedik bir hata oluştu: {e}")
        if "No tables found" in str(e):
             print("--- BİLGİ: Sitede 'No tables found' hatası alındı. Muhtemelen site güncelleniyor.")
        print("-" * 50)
    finally:
        is_running_lock.release()

# --- Yedekleme (Değişiklik yok) ---
def gunluk_ogleden_sonra_3_yedek():
    try:
        if not os.path.exists(EXCEL_DOSYASI):
            print(f"Yedek alınamadı: '{EXCEL_DOSYASI}' bulunamadı.")
            return
            
        tarih_str = datetime.now().strftime('%Y-%m-%d')
        hedef_klasor = os.path.join(YEDekLER_KLASORU, tarih_str)
        os.makedirs(hedef_klasor, exist_ok=True)
        
        saat_str = datetime.now().strftime('%H-%M-%S')
        hedef_dosya = os.path.join(hedef_klasor, f"kumluca_hal_fiyatlari_{saat_str}.xlsx")
        
        shutil.copy2(EXCEL_DOSYASI, hedef_dosya)
        print(f"--- BİLGİ: [Kumluca] Günlük yedek oluşturuldu: {hedef_dosya}")
    except Exception as e:
        print(f"!!! HATA: [Kumluca] Günlük yedek alınırken hata oluştu: {e}")

# --- Program Başlangıcı ve Zamanlama (Değişiklik yok) ---
if __name__ == '__main__':
    print("[Kumluca] Program başlatılıyor...")
    # support one-off execution
    if '--once' in sys.argv:
        verileri_cek_ve_kaydet()
        print("[Kumluca] --once ile tek çalışma tamamlandı. Çıkılıyor.")
        sys.exit(0)

    verileri_cek_ve_kaydet()
    schedule.every(15).minutes.do(verileri_cek_ve_kaydet)
    print("[Kumluca] Otomatik veri çekme işlemi her 15 dakikada bir zamanlandı.")

    schedule.every().day.at("15:00").do(gunluk_ogleden_sonra_3_yedek)
    print("[Kumluca] Günlük yedekleme işlemi her gün saat 15:00'e ayarlanmıştır.")
    print("Programı sonlandırmak için CTRL-C tuşlarına basın.")

    while True:
        schedule.run_pending()
        time.sleep(1)