import pandas as pd
import schedule
import time
from datetime import datetime
import ssl
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
import sys
import shutil
import threading
import urllib.error # <-- YENİDEN DENEME İÇİN EKLENDİ

# --- Kaynak Yolu (EXE için) ---
def kaynak_yolu(relative_path):
    """ .exe olarak paketlendiğinde doğru dosya yolunu bulur. """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# --- Kategori Kuralları ---
KATEGORI_DOSYASI = kaynak_yolu('kategoriler.xlsx')

def kategori_kural_yukle():
    if not os.path.exists(KATEGORI_DOSYASI):
        print(f"!!! HATA: '{KATEGORI_DOSYASI}' dosyası bulunamadı!")
        return None
    try:
        df = pd.read_excel(KATEGORI_DOSYASI)
        print(f"--- BİLGİ: '{KATEGORI_DOSYASI}' başarıyla yüklendi. İçinde {len(df)} kural bulundu.")
        if 'Anahtar_Kelime' not in df.columns or 'Kategori' not in df.columns:
            print("!!! HATA: 'kategoriler.xlsx' dosyasındaki sütun başlıkları yanlış!")
            return None
        return df
    except Exception as e:
        print(f"!!! HATA: '{KATEGORI_DOSYASI}' dosyası okunurken bir hata oluştu: {e}")
        return None

kategori_df = kategori_kural_yukle()

# !!!!!!!!! GÜNCEL: TÜRKÇE KARAKTER NORMALLEŞTİRME !!!!!!!!!
# Bu, "ÇARLİSTON BİBER" sorununu çözen DÜZELTİLMİŞ versiyondur.
def normalize_turkish(text):
    """
    Türkçe karakterleri (ve büyük/küçük harf) normalleştirir.
    .lower() fonksiyonunu en sonda çağırarak 'İ'/'ı' sorununu çözer.
    """
    if not isinstance(text, str):
        text = str(text)
        
    # Önce harf çevirme kuralları
    replacements = (
        ("ı", "i"), ("İ", "i"), # Önce 'ı' ve 'İ' harflerini 'i' yap
        ("ğ", "g"), ("Ğ", "g"),
        ("ü", "u"), ("Ü", "u"),
        ("ş", "s"), ("Ş", "s"),
        ("ö", "o"), ("Ö", "o"),
        ("ç", "c"), ("Ç", "c")
    )
    
    # \xa0 (non-breaking space) dahil, garip boşlukları temizle
    text = text.replace('\xa0', ' ')
    
    # Tüm harf çevirmelerini yap
    for old, new in replacements:
        text = text.replace(old, new)
        
    # EN SONUNDA tüm metni küçük harfe çevir
    return text.lower()

# !!!!!!!!! GÜNCEL: "Akıllı Eşleştirme" (BOŞLUK TEMİZLEMELİ) !!!!!!!!!
def kategori_belirle(urun_adi, kurallar_df):
    """
    Normalleştirme yaparak "akıllı" kategori belirler.
    'cılek', 'Çilek', 'BİBER' ve ' börülce ' (boşluklu) gibi farkları görmezden gelir.
    """
    if kurallar_df is None or kurallar_df.empty:
        return ''
    
    # 1. Siteden gelen ürün adını normalleştir
    urun_adi_normalized = normalize_turkish(urun_adi)
    
    for index, satir in kurallar_df.iterrows():
        
        # 2. Kuraldaki anahtar kelimeyi al, string'e çevir ve .strip() ile BOŞLUKLARI TEMİZLE
        anahtar_kelime_str = str(satir['Anahtar_Kelime']).strip() # <-- En son eklenen .strip()
        
        # 3. Boşlukları temizlenmiş kuralı normalleştir
        anahtar_kelime_normalized = normalize_turkish(anahtar_kelime_str)
        
        # 4. Normalleştirilmiş hallerini karşılaştır
        if anahtar_kelime_normalized in urun_adi_normalized:
            return satir['Kategori']
            
    return '' # Hiçbir kural eşleşmezse

# --- Global Ayarlar ---
ssl._create_default_https_context = ssl._create_unverified_context
BASE_URL = "https://eislem.izmir.bel.tr/tr/HalFiyatlari/20"
EXCEL_DOSYASI = "izmir_hal_fiyatlari.xlsx"
YEDekLER_KLASORU = "yedekler"

is_running_lock = threading.Lock()

# --- Excel Stilleri (Değişiklik yok) ---
def excel_stillerini_uygula(dosya_yolu):
    try:
        workbook = load_workbook(dosya_yolu)
        worksheet = workbook.active
        
        RENK_MAP = {
            "Meyve": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
            "Sebze": PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
            "Yeşillik": PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid"),
        }
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ince_kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        merkezi_hizalama = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet["1:1"]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = merkezi_hizalama; cell.border = ince_kenarlik

        for row_index in range(2, worksheet.max_row + 1):
            kategori_cell = worksheet.cell(row=row_index, column=2) # Kategori Sütunu (B)
            if kategori_cell.value:
                kategori = kategori_cell.value
                current_fill = RENK_MAP.get(kategori)

                for col_index in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_index, column=col_index)
                    cell.border = ince_kenarlik; cell.alignment = merkezi_hizalama
                    if current_fill: cell.fill = current_fill
        
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = max_length + 4
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
        
        workbook.save(dosya_yolu)
        print("--- BİLGİ: [İzmir] Gelişmiş Excel stilleri (Filtreleme dahil) başarıyla uygulandı.")
    except Exception as e:
        print(f"!!! HATA: [İzmir] Excel stilleri uygulanırken bir hata oluştu: {e}")

# --- Ana İşlem (Yeniden Deneme Mekanizması ile Güncellendi) ---
def verileri_cek_ve_kaydet():
    if kategori_df is None:
        print("Kategorizasyon kuralları yüklenemediği için işlem durduruldu.")
        return
        
    if not is_running_lock.acquire(blocking=False):
        print("Önceki görev tamamlanmadı; bu döngü atlandı.")
        return
        
    try:
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [İzmir] Görev başladı. Veriler çekiliyor...")
        
        # Dinamik Sütun Eşleştirme (Değişiklik yok)
        COLUMN_MAP = {
            'Ürün Adı': ['Adı', 'Mal Adı', 'Ürün Adı'],
            'Birim': ['Birimi', 'Birim'],
            'En Düşük Fiyat (TL)': ['En Az', 'En Az Fiyat', 'En Düşük Fiyat (TL)'],
            'En Yüksek Fiyat (TL)': ['En Çok', 'En Çok Fiyat', 'En Yüksek Fiyat (TL)']
        }
        STANDARD_COLS_ORDER = ['Ürün Adı', 'Kategori', 'En Düşük Fiyat (TL)', 'En Yüksek Fiyat (TL)', 'Birim']
        REQUIRED_STANDARD_COLS = list(COLUMN_MAP.keys()) 

        tarih_str = datetime.now().strftime("%Y-%m-%d")
        sebze_url = f"{BASE_URL}?date={tarih_str}&tip=1&aranacak="
        meyve_url = f"{BASE_URL}?date={tarih_str}&tip=2&aranacak="
        
        valid_dfs = [] 

        for url_type, url in [("Sebze", sebze_url), ("Meyve", meyve_url)]:
            
            # Yeniden Deneme Bloğu (Değişiklik yok)
            df_list = None
            max_retries = 3
            for attempt in range(1, max_retries + 1):
                try:
                    print(f"--- BİLGİ: [İzmir] {url_type} verisi çekiliyor (Deneme {attempt}/{max_retries}): {url}")
                    df_list = pd.read_html(url) # AĞ İSTEĞİ
                    
                    print(f"--- BİLGİ: [İzmir] {url_type} verisi (Deneme {attempt}/{max_retries}) BAŞARILI.")
                    break 
                    
                except (urllib.error.URLError, ssl.SSLError, ConnectionError, OSError) as e:
                    print(f"--- UYARI: [İzmir] {url_type} AĞ HATASI (Deneme {attempt}/{max_retries}): {e}")
                    if attempt < max_retries:
                        print(f"    ... 3 saniye beklenip yeniden denenecek ...")
                        time.sleep(3)
                    else:
                        print(f"    ... {max_retries} deneme de başarısız oldu. Bu URL atlanıyor.")
                except Exception as e:
                    print(f"--- UYARI: [İzmir] {url_type} VERİ HATASI (muhtemelen boş sayfa/veri yok): {e}")
                    break
            
            # Yeniden Deneme Bloğu Sonu (Değişiklik yok)
            
            if not df_list:
                print(f"--- UYARI: [İzmir] {url_type} verisi {max_retries} deneme sonunda alınamadı/bulunamadı.")
                continue 
            
            # Dinamik Eşleştirme Bloğu (Değişiklik yok)
            try:
                df = df_list[0]
                actual_columns = list(df.columns) 
                
                rename_map = {} 
                found_cols = [] 
                
                for standard_name, possible_names in COLUMN_MAP.items():
                    for possible in possible_names:
                        if possible in actual_columns:
                            rename_map[possible] = standard_name
                            found_cols.append(standard_name)
                            break 
                
                if all(col in found_cols for col in REQUIRED_STANDARD_COLS):
                    df_renamed = df.rename(columns=rename_map)
                    df_final = df_renamed[REQUIRED_STANDARD_COLS] 
                    valid_dfs.append(df_final)
                    print(f"--- BİLGİ: [İzmir] {url_type} verisi bulundu ve {rename_map} ile eşleştirildi.")
                else:
                    print(f"--- UYARI: [İzmir] {url_type} tablosunda gerekli sütunlar bulunamadı.")
                    print(f"    Siteden gelen: {actual_columns}")
                    print(f"    Bulunabilenler: {found_cols}")
            
            except Exception as e:
                print(f"!!! HATA: [İzmir] {url_type} verisi işlenirken (eşleştirme) hata oluştu: {e}")
        
        # --- URL Döngüsü Sonu ---
        
        if not valid_dfs:
            print("!!! HATA: [İzmir] Bugün için geçerli Sebze veya Meyve verisi bulunamadı. İşlem atlanıyor.")
            return 

        toplam_df = pd.concat(valid_dfs, ignore_index=True)
        fiyat_df = toplam_df

        # Bu satır artık "börülce", "darı", "biber" ve "boşluk" sorunlarını
        # çözen en güncel 'kategori_belirle' fonksiyonunu kullanıyor.
        fiyat_df['Kategori'] = fiyat_df['Ürün Adı'].apply(lambda urun: kategori_belirle(urun, kategori_df))
        fiyat_df = fiyat_df[STANDARD_COLS_ORDER] 
        
        print("--- BİLGİ: [İzmir] Kategorizasyon sonrası verilerin ilk 5 satırı:")
        print(fiyat_df.head())
        
        fiyat_df.to_excel(EXCEL_DOSYASI, index=False, engine='openpyxl')
        excel_stillerini_uygula(EXCEL_DOSYASI)
        
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [İzmir] Veriler başarıyla '{EXCEL_DOSYASI}' dosyasına kaydedildi.")
        print("-" * 50)
        
    except Exception as e:
        print(f"!!! HATA: [İzmir] Ana işlem sırasında beklenmedik bir hata oluştu: {e}")
        print("-" * 50)
    finally:
        is_running_lock.release()

# --- Yedekleme (Değişiklik yok) ---
def gunluk_ogleden_sonra_3_yedek():
    try:
        if not os.path.exists(EXCEL_DOSYASI):
            print(f"Yedek alınamadı: '{EXCEL_DOSYASI}' bulunamadı.")
            return
            
        tarih_str = datetime.now().strftime('%Y-%m-%d')
        hedef_klasor = os.path.join(YEDekLER_KLASORU, tarih_str)
        os.makedirs(hedef_klasor, exist_ok=True)
        
        saat_str = datetime.now().strftime('%H-%M-%S')
        hedef_dosya = os.path.join(hedef_klasor, f"izmir_hal_fiyatlari_{saat_str}.xlsx")
        
        shutil.copy2(EXCEL_DOSYASI, hedef_dosya)
        print(f"--- BİLGİ: [İzmir] Günlük yedek oluşturuldu: {hedef_dosya}")
    except Exception as e:
        print(f"!!! HATA: [İzmir] Günlük yedek alınırken hata oluştu: {e}")

# --- Program Başlangıcı ve Zamanlama (Değişiklik yok) ---
if __name__ == '__main__':
    print("[İzmir] Program başlatılıyor...")
    # support one-off execution
    if '--once' in sys.argv:
        verileri_cek_ve_kaydet()
        print("[İzmir] --once ile tek çalışma tamamlandı. Çıkılıyor.")
        sys.exit(0)

    verileri_cek_ve_kaydet()
    schedule.every(15).minutes.do(verileri_cek_ve_kaydet)
    print("[İzmir] Otomatik veri çekme işlemi her 15 dakikada bir zamanlandı.")

    schedule.every().day.at("15:00").do(gunluk_ogleden_sonra_3_yedek)
    print("[İzmir] Günlük yedekleme işlemi her gün saat 15:00'e ayarlanmıştır.")
    print("Programı sonlandırmak için CTRL-C tuşlarına basın.")

    while True:
        schedule.run_pending()
        time.sleep(1)