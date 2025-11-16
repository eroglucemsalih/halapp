import pandas as pd
import schedule
import time
from datetime import datetime, timedelta
import ssl
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
import sys
import shutil
import threading
import urllib.error
import requests
from bs4 import BeautifulSoup
import urllib3
import re
import numpy as np
import logging
from logging.handlers import RotatingFileHandler
from contextlib import ExitStack
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

# SSL/TLS sertifika uyarılarını (InsecureRequestWarning) kapat
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- Loglama Ayarları ---
log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - [Gazipaşa] %(message)s')
log_file = "gazipasa_hal.log"
log_handler = RotatingFileHandler(log_file, maxBytes=1024*1024, backupCount=5)
log_handler.setFormatter(log_formatter)

logger = logging.getLogger('GazipasaHal')
logger.setLevel(logging.INFO)
# Konsola da log basmak için:
logger.addHandler(logging.StreamHandler(sys.stdout))
logger.addHandler(log_handler)

# --- Cache Yönetimi ---
def cache_son_veri(df, cache_file="son_basarili_veri.pkl"):
    """Son başarılı veriyi pickle olarak sakla"""
    try:
        df.to_pickle(cache_file)
        logger.info("Son başarılı veri cache'lendi")
    except Exception as e:
        logger.error(f"Cache kaydetme hatası: {e}")

def son_cache_getir(cache_file="son_basarili_veri.pkl"):
    """Son başarılı veriyi getir"""
    try:
        if os.path.exists(cache_file):
            return pd.read_pickle(cache_file)
    except Exception as e:
        logger.error(f"Cache okuma hatası: {e}")
    return None

# --- Retry Decorator ---
# Sadece spesifik ağ hataları için yeniden dene
RETRY_EXCEPTIONS = (
    urllib.error.URLError, 
    ssl.SSLError, 
    ConnectionError, 
    OSError, 
    requests.exceptions.RequestException
)
@retry(stop=stop_after_attempt(3),
       wait=wait_exponential(multiplier=1, min=2, max=10),
       retry=retry_if_exception_type(RETRY_EXCEPTIONS))
def fetch_data_with_retry(session, url):
    """Exponential backoff ile retry mekanizması"""
    logger.info(f"Fetch denemesi: {url}")
    response = session.get(url, headers={'User-Agent': 'Mozilla/5.0'}, verify=False, timeout=10)
    response.raise_for_status()
    return response

# --- Kaynak Yolu (EXE için) ---
def kaynak_yolu(relative_path):
    """ .exe olarak paketlendiğinde doğru dosya yolunu bulur. """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# --- Kategori Kuralları ---
KATEGORI_DOSYASI = kaynak_yolu('kategoriler.xlsx')

def kategori_kural_yukle():
    if not os.path.exists(KATEGORI_DOSYASI):
        logger.error(f"'{KATEGORI_DOSYASI}' dosyası bulunamadı!")
        return None
    try:
        df = pd.read_excel(KATEGORI_DOSYASI)
        logger.info(f"'{KATEGORI_DOSYASI}' başarıyla yüklendi. İçinde {len(df)} kural bulundu.")
        if 'Anahtar_Kelime' not in df.columns or 'Kategori' not in df.columns:
            logger.error("'kategoriler.xlsx' dosyasındaki sütun başlıkları yanlış!")
            return None
        return df
    except Exception as e:
        logger.error(f"'{KATEGORI_DOSYASI}' dosyası okunurken bir hata oluştu: {e}")
        return None

kategori_df = kategori_kural_yukle()

# --- Akıllı Kategorizasyon (En son 'strip'li ve '\xa0' düzeltmeli) ---
def normalize_turkish(text):
    if not isinstance(text, str):
        text = str(text)
    replacements = (
        ("ı", "i"), ("İ", "i"), ("ğ", "g"), ("Ğ", "g"), ("ü", "u"), ("Ü", "u"),
        ("ş", "s"), ("Ş", "s"), ("ö", "o"), ("Ö", "o"), ("ç", "c"), ("Ç", "c")
    )
    # \xa0 (non-breaking space) dahil, garip boşlukları temizle
    text = text.replace('\xa0', ' ')
    
    for old, new in replacements:
        text = text.replace(old, new)
    return text.lower()

def kategori_belirle(urun_adi, kurallar_df):
    if kurallar_df is None or kurallar_df.empty:
        return ''
    urun_adi_normalized = normalize_turkish(urun_adi)
    for index, satir in kurallar_df.iterrows():
        anahtar_kelime_str = str(satir['Anahtar_Kelime']).strip()
        anahtar_kelime_normalized = normalize_turkish(anahtar_kelime_str)
        if anahtar_kelime_normalized in urun_adi_normalized:
            return satir['Kategori']
    return ''

# --- URL İşleme Yardımcıları ---
def is_valid_link(link, keywords, block_words):
    """Link geçerlilik kontrolü"""
    link_text = normalize_turkish(link.get_text().strip())
    return (any(re.search(keyword, link_text) for keyword in keywords) and
            all(not re.search(block_word, link_text) for block_word in block_words))

def build_full_url(href, base_domain):
    """URL'yi tam haline getir"""
    if href.startswith('http'):
        return href
    elif href.startswith('//'):
        return 'https:' + href
    elif href.startswith('/'):
        return base_domain + href
    return None

def find_data_url(session, base_url, link_keywords, block_keywords, base_domain):
    """Daha modüler link bulma fonksiyonu"""
    try:
        response = fetch_data_with_retry(session, base_url)
        
        soup = BeautifulSoup(response.content, 'html.parser')
        all_links = soup.find_all('a', href=True)
        
        for link in all_links:
            href = link.get('href')
            if not href:
                continue
                
            if is_valid_link(link, link_keywords, block_keywords):
                full_url = build_full_url(href, base_domain)
                if full_url and full_url != base_url:
                    logger.info(f"Dinamik link bulundu: {full_url}")
                    return full_url
                    
    except Exception as e:
        logger.error(f"Link bulma hatası: {e}")
    return None

# --- Tablo İşleme Yardımcıları ---
def clean_price_column(series):
    """Fiyat sütunu temizleme"""
    return (series.astype(str)
            .str.replace('₺', '', regex=False)
            .str.replace(',', '.', regex=False)
            .str.strip())

# !!!!!!!!! GÜNCELLENDİ: 'standardize_table' (Artık sadece temizler) !!!!!!!!!
def standardize_table(df):
    """Tablo standartlaştırma (YENİDEN ADLANDIRILMIŞ DF bekler)"""
    # Fiyatları temizle
    for col in ['En Düşük Fiyat (TL)', 'En Yüksek Fiyat (TL)']:
        df[col] = clean_price_column(df[col])
    
    # Boş/kirli verileri (ara başlıklar) np.nan yap
    df.replace([r'^\s*$', r'\*\*', 'nan', 'None'], np.nan, regex=True, inplace=True)
    
    # "Gereksiz verileri" (ara başlıklar) filtrele
    # 'Grup' sütunu NaN olan (yani ara başlık olan) satırları at
    df = df.dropna(subset=['Grup'])
    
    df['Birim'] = 'KG'  # PDF'e [kaynak: 7] göre sabit
    return df

def find_and_process_table(tablolar, fingerprint="toptanci hal müdürlüğü"):
    """Daha modüler tablo bulma ve işleme"""
    df_raw = None
    
    for table in tablolar:
        if table.empty:
            continue
        if fingerprint in str(table.columns).lower():
            df_raw = table.copy()
            logger.info("Ana veri tablosu ('TOPTANCI HAL...') [kaynak: 3] bulundu.")
            break
            
    if df_raw is None:
        logger.error("İşlenecek 'TOPTANCI HAL...' [kaynak: 3] tablosu bulunamadı")
        return None

    # "ÜRÜN ADI" [kaynak: 4, 7] başlıklarını bul
    header_indices = []
    for i, cell in enumerate(df_raw.iloc[:, 0]): # Sadece ilk sütuna bak
        if "ÜRÜN ADI" in str(cell).upper():
            header_indices.append(i)

    if not header_indices:
        logger.error("'ÜRÜN ADI' [kaynak: 4, 7] başlığı bulunamadı")
        return None
        
    # Sadece SON (Dernek) tabloyu al
    dernek_start_index = header_indices[-1]
    df_dernek_chunk = df_raw.iloc[dernek_start_index:]
    
    # Başlığı ayarla
    new_headers = df_dernek_chunk.iloc[0]
    df_clean = df_dernek_chunk[1:]
    df_clean.columns = new_headers
    df_clean = df_clean.reset_index(drop=True)
    
    logger.info(f"Son 'ÜRÜN ADI' parçası (Dernek Tablosu) {len(df_clean)} satırla ayrıldı.")
    return df_clean

# --- Global Ayarlar ---
ssl._create_default_https_context = ssl._create_unverified_context
URL_LISTING = "https://gazipasa.bel.tr/gunluk-hal-fiyatlari" 
BASE_DOMAIN = "https://gazipasa.bel.tr" 
EXCEL_DOSYASI = "gazipasa_hal_fiyatlari.xlsx"
YEDekLER_KLASORU = "yedekler"

is_running_lock = threading.Lock()

# --- Excel Stilleri (Tek sayfaya uygular) ---
def apply_styling_to_sheet(worksheet):
    """Verilen openpyxl worksheet nesnesine stilleri uygular."""
    try:
        RENK_MAP = {
            "Meyve": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
            "Sebze": PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
            "Yeşillik": PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid"),
        }
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ince_kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        merkezi_hizalama = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # !!!!!!!!! YENİ: Sütun sırası değiştiği için Kategori sütunu 3. (C) sütun !!!!!!!!!
        KATEGORI_SUTUNU = 3 

        for cell in worksheet["1:1"]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = merkezi_hizalama; cell.border = ince_kenarlik

        for row_index in range(2, worksheet.max_row + 1):
            kategori_cell = worksheet.cell(row=row_index, column=KATEGORI_SUTUNU)
            if kategori_cell.value:
                kategori = kategori_cell.value
                current_fill = RENK_MAP.get(kategori)

                for col_index in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_index, column=col_index)
                    cell.border = ince_kenarlik; cell.alignment = merkezi_hizalama
                    if current_fill: cell.fill = current_fill
        
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = max_length + 4
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
        
        logger.info(f"'{worksheet.title}' sayfası için stiller uygulandı.")
    except Exception as e:
        logger.error(f"'{worksheet.title}' stilleri uygulanırken bir hata oluştu: {e}")

# --- Ana İşlem (İyileştirilmiş) ---
def check_prerequisites():
    """Ön koşulları kontrol et"""
    if kategori_df is None:
        logger.error("Kategorizasyon kuralları yüklenemediği için işlem durduruldu")
        return False
    return True

# !!!!!!!!! GÜNCELLENDİ: Ana İşlem (Mantık sırası ve GRUP sütunu düzeltildi) !!!!!!!!!
def verileri_cek_ve_kaydet():
    """Ana veri çekme ve işleme fonksiyonu - geliştirilmiş"""
    if not check_prerequisites():
        return

    with ExitStack() as stack:
        if not is_running_lock.acquire(blocking=False):
            logger.warning("Önceki görev tamamlanmadı; bu döngü atlandı")
            return
        stack.callback(is_running_lock.release) # Görev bitince kilidi aç
        
        try:
            logger.info("Görev başladı, veriler çekiliyor...")
            
            # --- Sabit Konfigürasyon ---
            # TESPİT'e göre (log 21:41:03) Dernek tablosunun haritası
            COLUMN_MAP = {
                'Ürün Adı': ['ÜRÜN ADI'],
                'Birim': ['BİRİMİ (KG)'],
                'En Düşük Fiyat (TL)': ['FİYAT (TL)'],
                'En Yüksek Fiyat (TL)': ['FİYAT (TL).1']
            }
            # !!!!!!!!! YENİ: Sütun sırasına "Grup" eklendi !!!!!!!!!
            STANDARD_COLS_ORDER = ['Grup', 'Ürün Adı', 'Kategori', 'En Düşük Fiyat (TL)', 'En Yüksek Fiyat (TL)', 'Birim']
            REQUIRED_STANDARD_COLS = list(COLUMN_MAP.keys())
            
            LINK_KEYWORDS = ["fiyatları", r"\bhal\b"]
            LINK_BLOCK_KEYWORDS = [r"\bihale\b", r"\bhalk\b"]
            
            # --- Veri Çekme ve İşleme ---
            session = requests.Session()
            
            # 1. Dinamik link bul
            data_url = find_data_url(session, URL_LISTING, LINK_KEYWORDS, LINK_BLOCK_KEYWORDS, BASE_DOMAIN)
            if not data_url:
                raise ValueError("Güncel veri linki bulunamadı")
                
            # 2. Veriyi çek (retry mekanizması ile)
            response = fetch_data_with_retry(session, data_url)
            # header=0 -> Log'lara göre bu, 'TOPTANCI HAL...' [kaynak: 3] başlığını bulan doğru parametre
            tablolar = pd.read_html(response.content, header=0)
            logger.info(f"{len(tablolar)} tablo bulundu")
                
            # 3. Tabloları işle (Sadece Dernek tablosunu al)
            df_clean = find_and_process_table(tablolar)
            if df_clean is None:
                raise ValueError("Dernek tablosu işlenemedi")
            
            # !!!!!!!!! DÜZELTME BAŞLANGICI: MANTIK SIRASI DÜZELTİLDİ !!!!!!!!!
            
            # 4. YENİDEN ADLANDIR (Hata buradaydı)
            # Pandas'ın bozuk okumasını düzelt (FİYAT (TL), FİYAT (TL).1)
            df_clean.columns = pd.io.common.dedup_names(df_clean.columns, is_potential_multiindex=False)
            actual_columns_normalized = [normalize_turkish(str(col)).strip() for col in df_clean.columns]
            
            rename_map = {}
            found_cols = []
            
            for standard_name, possible_names_list in COLUMN_MAP.items():
                for possible_name in possible_names_list: 
                    possible_normalized = normalize_turkish(possible_name).strip()
                    
                    if possible_normalized in actual_columns_normalized:
                        real_col_name = df_clean.columns[actual_columns_normalized.index(possible_normalized)]
                        rename_map[real_col_name] = standard_name
                        found_cols.append(standard_name)
                        actual_columns_normalized[actual_columns_normalized.index(possible_normalized)] = "USED" 
                        break 
                if standard_name in found_cols:
                    continue 
            
            if not all(col in found_cols for col in REQUIRED_STANDARD_COLS):
                logger.error("Dernek tablosu bulundu ama sütunlar eşleşmedi.")
                logger.error(f"ARANAN (normalize): {[normalize_turkish(p[0]) for p_list in COLUMN_MAP.values() for p in p_list]}")
                logger.error(f"BULUNAN (normalize): {[normalize_turkish(str(col)).strip() for col in df_clean.columns]}")
                raise ValueError("Sütun eşleştirme başarısız")
            
            df_renamed = df_clean.rename(columns=rename_map)
            
            # !!!!!!!!! YENİ ADIM: "Grup" Sütununu Oluşturma (İsteğiniz) !!!!!!!!!
            
            # Önce fiyat sütunlarını (temizlikten önce) kopyala
            fiyat_sutun_1 = clean_price_column(df_renamed['En Düşük Fiyat (TL)'])
            fiyat_sutun_2 = clean_price_column(df_renamed['En Yüksek Fiyat (TL)'])
            
            groups = []
            current_group = 'Diğer' # Varsayılan grup
            
            for index, row in df_renamed.iterrows():
                urun_adi = str(row.get('Ürün Adı', ''))
                
                # Fiyat sütunları boş mu diye (temizlenmiş hallerine) bak
                # Boş string '', 'nan', 'None' veya '**' olabilir, hepsi pd.isna() ile temizlenir
                fiyat_1_bos_mu = pd.isna(pd.to_numeric(fiyat_sutun_1.iloc[index], errors='coerce'))
                fiyat_2_bos_mu = pd.isna(pd.to_numeric(fiyat_sutun_2.iloc[index], errors='coerce'))
                
                # Eğer fiyatlar boşsa VE ürün adında -LAR/-LER varsa, bu bir ara başlıktır [kaynak: 4, 7]
                if (fiyat_1_bos_mu and fiyat_2_bos_mu) and ('LAR' in urun_adi.upper() or 'LER' in urun_adi.upper()):
                    current_group = urun_adi # Yeni grubu ayarla
                    groups.append(np.nan) # Bu satır gereksiz, silinmek üzere işaretle
                else:
                    groups.append(current_group) # Bu bir veri satırı, mevcut grubu ata

            df_renamed['Grup'] = groups
            logger.info("Grup sütunu oluşturuldu.")

            # 5. Tabloyu standartlaştır (Artık YENİDEN ADLANDIRILMIŞ tabloyu kullanır)
            fiyat_df = standardize_table(df_renamed) 
            
            # !!!!!!!!! DÜZELTME SONU !!!!!!!!!
            
            # 6. Eksik hücreleri doldurma ve '-lar'/'-ler' ile biten satırları temizleme
            # Önce tüm eksik hücrelere "veri yok" değeri ata (string olarak)
            fiyat_df = fiyat_df.fillna('veri yok')

            # Satır bazında: eğer bir satırda 3 veya daha fazla hücrede "veri yok" varsa o satırı sil
            # (DİKKAT: satır bazlı kontrol, sütun bazlı değil)
            veri_nok_sayisi = (fiyat_df.astype(str) == 'veri yok').sum(axis=1)
            remove_rows_mask = veri_nok_sayisi >= 3
            if remove_rows_mask.any():
                logger.info(f"{remove_rows_mask.sum()} satırta >=3 'veri yok' bulundu; bu satırlar siliniyor.")
                fiyat_df = fiyat_df[~remove_rows_mask].reset_index(drop=True)

            # 'Ürün Adı' sütununda '-lar' veya '-ler' ile biten satırları (büyük/küçük harf duyarsız)
            # tamamen kaldır. Bu, ara başlıkların artık tabloda kalmaması içindir.
            urun_adi_series = fiyat_df['Ürün Adı'].astype(str).str.strip().str.lower()
            remove_mask = urun_adi_series.str.endswith(('lar', 'ler'))
            if remove_mask.any():
                logger.info(f"{remove_mask.sum()} satır '-lar'/'-ler' eki nedeniyle siliniyor.")
                fiyat_df = fiyat_df[~remove_mask].reset_index(drop=True)

            # 7. Kategorizasyon uygula
            fiyat_df['Kategori'] = fiyat_df['Ürün Adı'].apply(lambda urun: kategori_belirle(urun, kategori_df))
            fiyat_df = fiyat_df[STANDARD_COLS_ORDER] # Yeni sıralamayı uygula
            
            logger.info(f"'Dernek' tablosu başarıyla işlendi. {len(fiyat_df)} temiz ürün bulundu.")
            
            # 7. Excel'e kaydet ve cache'le
            with pd.ExcelWriter(EXCEL_DOSYASI, engine='openpyxl') as writer:
                fiyat_df.to_excel(writer, sheet_name='Hal_Fiyatlari', index=False)
                workbook = writer.book
                apply_styling_to_sheet(workbook['Hal_Fiyatlari'])
                
            # Son başarılı veriyi cache'le
            cache_son_veri(fiyat_df)
            logger.info(f"Veriler başarıyla '{EXCEL_DOSYASI}' dosyasına kaydedildi")
            print(f"--- BİLGİ: [Gazipaşa] Kategorizasyon sonrası verilerin ilk 5 satırı:") # Konsola da basalım
            print(fiyat_df.head())
            return fiyat_df
            
        except Exception as e:
            logger.error(f"Ana işlem sırasında beklenmedik bir hata oluştu: {e}")
            if "No tables found" in str(e):
                logger.info("Sitede 'No tables found' hatası alındı. Muhtemelen site güncelleniyor.")
            
            # Hata durumunda cache'i yükle
            cached_data = son_cache_getir()
            if cached_data is not None:
                logger.info("Hata nedeniyle son başarılı veri cache'den yüklendi")
                return cached_data
            else:
                logger.error("Cache'de veri yok, görev başarısız oldu.")
                raise # Eğer cache de yoksa hatayı scheduler'a fırlat
        
# --- Yedekleme İyileştirilmiş ---
def cleanup_old_backups(backup_dir, days=30):
    """Eski yedekleri temizle"""
    try:
        cutoff_date = datetime.now() - timedelta(days=days)
        for root, dirs, files in os.walk(backup_dir):
            for dir_name in dirs:
                try:
                    dir_date = datetime.strptime(dir_name, '%Y-%m-%d')
                    if dir_date < cutoff_date:
                        dir_path = os.path.join(root, dir_name)
                        shutil.rmtree(dir_path)
                        logger.info(f"Eski yedek klasörü silindi: {dir_path}")
                except ValueError:
                    continue # Tarih formatına uymayan klasörler
    except Exception as e:
        logger.error(f"Eski yedek temizleme hatası: {e}")

def gunluk_ogleden_sonra_3_yedek():
    """Günlük yedekleme - geliştirilmiş versiyon"""
    try:
        if not os.path.exists(EXCEL_DOSYASI):
            logger.error(f"Yedek alınamadı: '{EXCEL_DOSYASI}' bulunamadı")
            return
            
        # Yedek klasörü oluştur
        tarih_str = datetime.now().strftime('%Y-%m-%d')
        hedef_klasor = os.path.join(YEDekLER_KLASORU, tarih_str)
        os.makedirs(hedef_klasor, exist_ok=True)
        
        # Yedek dosyasını oluştur
        saat_str = datetime.now().strftime('%H-%M-%S')
        hedef_dosya = os.path.join(hedef_klasor, f"gazipasa_hal_fiyatlari_{saat_str}.xlsx")
        
        # Yedek al ve eski yedekleri temizle
        shutil.copy2(EXCEL_DOSYASI, hedef_dosya)
        cleanup_old_backups(YEDekLER_KLASORU)
        
        logger.info(f"Günlük yedek oluşturuldu: {hedef_dosya}")
    except Exception as e:
        logger.error(f"Günlük yedek alınırken hata oluştu: {e}")

# --- Program Başlangıcı ve Zamanlama (Değişiklik yok) ---
if __name__ == '__main__':
    print("[Gazipaşa] Program başlatılıyor...")
    # support one-off execution
    if '--once' in sys.argv:
        verileri_cek_ve_kaydet()
        print("[Gazipaşa] --once ile tek çalışma tamamlandı. Çıkılıyor.")
        sys.exit(0)

    verileri_cek_ve_kaydet()
    schedule.every(15).minutes.do(verileri_cek_ve_kaydet)
    print("[Gazipaşa] Otomatik veri çekme işlemi her 15 dakikada bir zamanlandı.")

    schedule.every().day.at("15:00").do(gunluk_ogleden_sonra_3_yedek)
    print("[Gazipaşa] Günlük yedekleme işlemi her gün saat 15:00'e ayarlanmıştır.")
    print("Programı sonlandırmak için CTRL-C tuşlarına basın.")

    while True:
        schedule.run_pending()
        time.sleep(1)